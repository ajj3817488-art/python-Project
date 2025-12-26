import os, datetime, csv, shutil
import openpyxl
from openpyxl import Workbook
from zipfile import BadZipFile
from tkinter import *
from tkinter import ttk, messagebox, filedialog, Toplevel
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
import arabic_reshaper
from bidi.algorithm import get_display
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# ====== إنشاء ملفات Excel إذا لم تكن موجودة ======
def create_users_excel():
    try:
        wb = openpyxl.load_workbook("users.xlsx")
    except (FileNotFoundError, BadZipFile):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["Username", "Password", "Role"])
        ws.append(["admin", "1234", "Admin"])
        ws.append(["seller", "1111", "Seller"])
        wb.save("users.xlsx")

def create_products_excel():
    try:
        wb = openpyxl.load_workbook("products.xlsx")
    except (FileNotFoundError, BadZipFile):
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Product Name", "Price", "Quantity", "Category", "Image Path"])
        wb.save("products.xlsx")

def create_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "customer"
    ws.append(["Full Name", "Phone", "Address", "Total", "Date"])
    wb.save("raken.xlsx")

# تشغيلها مرة واحدة عند بدء البرنامج
create_users_excel()
create_products_excel()
try:
    wb = openpyxl.load_workbook("raken.xlsx")
except (FileNotFoundError, BadZipFile):
    create_excel()

# ====== دوال مساعدة ======
def arabic_text(text):
    reshaped_text = arabic_reshaper.reshape(text)
    return get_display(reshaped_text)

# ====== شاشة تسجيل الدخول ======
def login_screen():
    login_win = Tk()
    login_win.title("🔐 تسجيل الدخول")
    login_win.geometry("400x300")
    login_win.configure(bg="#EFEAD8")
    login_win.resizable(False, False)

    Label(login_win, text="🧾 نظام المبيعات - تسجيل الدخول", bg="#EFEAD8",
          font=("Tajawal", 14, "bold")).pack(pady=20)
    Label(login_win, text="اسم المستخدم:", bg="#EFEAD8").pack(pady=5)
    user_entry = Entry(login_win, width=30, font=("Tajawal", 12))
    user_entry.pack()
    Label(login_win, text="كلمة المرور:", bg="#EFEAD8").pack(pady=5)
    pass_entry = Entry(login_win, width=30, font=("Tajawal", 12), show="*")
    pass_entry.pack()

    def verify_login():
        username = user_entry.get().strip()
        password = pass_entry.get().strip()
        if username == "" or password == "":
            messagebox.showwarning("تنبيه", "الرجاء إدخال اسم المستخدم وكلمة المرور.")
            return
        try:
            wb = openpyxl.load_workbook("users.xlsx")
            ws = wb.active
        except:
            create_users_excel()
            wb = openpyxl.load_workbook("users.xlsx")
            ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            user, pw, role = row
            if username == user and password == pw:
                messagebox.showinfo("تم", f"✅ مرحبًا {username} ({role})")
                login_win.destroy()
                open_main_window(role)
                return
        messagebox.showerror("خطأ", "❌ اسم المستخدم أو كلمة المرور غير صحيحة.")

    Button(login_win, text="تسجيل الدخول", bg="#6D8B74", fg="white", width=20,
           font=("Tajawal", 12), command=verify_login).pack(pady=15)
    login_win.mainloop()

# ====== البرنامج الرئيسي ======
def open_main_window(role):
    global SA
    SA = Tk()
    SA.geometry("1200x600")
    SA.title(f"Market tools for Building [ملابس] - {role}")

    if role == "Seller":
        manage_btn_state = DISABLED
        report_btn_state = DISABLED
    else:
        manage_btn_state = NORMAL
        report_btn_state = NORMAL

    sa = Frame(SA, bg='silver', width=600, height=550)
    sa.place(x=1, y=1)

    title = Label(sa, text=f"مرحبًا {role}", font=("Tajawal", 13), fg="white", bg="#5F7161", width=70)
    title.place(x=0, y=0)

    # قائمة المنتجات الافتراضية
    global menu, sb, F2, F3, F4, F6, F8, hj, tax_entry, discount_entry
    menu = {}  # سيتم تعبئتها من products.xlsx
    sb = []

    # ====== تحميل المنتجات من Excel ======
    try:
        wb_products = openpyxl.load_workbook("products.xlsx")
        ws_products = wb_products.active
        for i, row in enumerate(ws_products.iter_rows(min_row=2, values_only=True)):
            name, price, qty, cat, img_path = row
            menu[i] = [name, float(price), int(qty), cat, img_path]
    except:
        pass

    # ====== إنشاء Spinbox لكل منتج ======
    for i in range(len(menu)):
        var = IntVar()
        spin = Spinbox(SA, from_=0, to_=20, font=("Times", 12), width=10, textvariable=var)
        spin.place(x=30 + i*120, y=140)
        sb.append(spin)

    # ====== واجهة عرض المنتجات مع الصور ======
    for i, data in menu.items():
        name, price, qty, cat, img_path = data
        if img_path and os.path.exists(img_path):
            img = PhotoImage(file=img_path)
        else:
            img = None
        btn = Button(sa, width=88, height=85, bg="#EFEAD8", bd=1, relief=SOLID,
                     text=name, image=img, compound=TOP)
        btn.image = img  # للحفاظ على المرجع
        btn.place(x=30 + i*120, y=45)

    # ====== شجرة الفاتورة ======
    SA1 = Frame(SA, bg="gray", width=343, height=550)
    SA1.place(x=604, y=1)
    hj = ttk.Treeview(SA1, selectmode="browse")
    hj.place(x=1, y=1, width=340, height=550)
    hj["columns"] = ("1", "2", "3")
    hj.column("#0", width=80, anchor="c")
    hj.column("1", width=50, anchor="c")
    hj.column("2", width=50, anchor="c")
    hj.column("3", width=60, anchor="c")
    hj.heading("#0", text="المواد", anchor="c")
    hj.heading("1", text="السعر", anchor="c")
    hj.heading("2", text="العدد", anchor="c")
    hj.heading("3", text="الإجمالي", anchor="c")

    # ====== أزرار أسفل الشاشة ======
    Button(SA, text="🛒 شراء", fg="white", font=("Tajawal", 12),
           width=15, bg="#6D8B74", bd=1, relief=SOLID, cursor="hand2", height=1,
           command=bill).place(x=30, y=500)
    Button(SA, text="🧾 فاتورة جديدة", fg="white", font=("Tajawal", 12),
           width=15, bg="#6D8B74", bd=1, relief=SOLID, cursor="hand2", height=1,
           command=clear).place(x=180, y=500)
    Button(SA, text ="قاىمه العملاء", fg="white", font=("Tajawal", 12),
           width=15, bg="#6D8B74", bd=1, relief=SOLID, cursor="hand2", height=1,
           command=show_all_invoices).place(x=330, y=500)
    Button(SA, text ="اداره المنتجات", fg="white", font=("Tajawal", 12),
           width=14, bg="#6D8B74", bd=1, relief=SOLID, cursor="hand2", height=1,
           state=manage_btn_state, command=manage_products).place(x=475, y=500)
    Button(SA, text="📊 التقارير", fg="white", font=("Tajawal", 12),
           width=15, bg="#5F7161", bd=1, relief=SOLID, cursor="hand2", height=1,
           state=report_btn_state, command=show_dashboard).place(x=620, y=500)

    # زر تسجيل الخروج
    Button(SA, text="🚪 تسجيل الخروج", bg="#C65D7B", fg="white", font=("Tajawal", 11),
           width=15, command=lambda: (SA.destroy(), login_screen())).place(x=780, y=10)

    SA.mainloop()

# ====== إدارة المنتجات مع إضافة الصور ======
def manage_products():
    win = Toplevel(SA)
    win.title("📦 إدارة المنتجات")
    win.geometry("900x500")
    win.configure(bg="#F8F6F0")
    win.resizable(False, False)

    try:
        wb = openpyxl.load_workbook("products.xlsx")
    except:
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Product Name", "Price", "Quantity", "Category", "Image Path"])
        wb.save("products.xlsx")
        wb = openpyxl.load_workbook("products.xlsx")
    ws = wb.active

    # ====== قسم الإدخال ======
    frame_top = Frame(win, bg="#F8F6F0")
    frame_top.pack(pady=10)
    Label(frame_top, text="اسم المنتج:", bg="#F8F6F0", font=("Tajawal", 11)).grid(row=0, column=0, padx=5)
    name_entry = Entry(frame_top, width=15)
    name_entry.grid(row=0, column=1, padx=5)
    Label(frame_top, text="السعر:", bg="#F8F6F0", font=("Tajawal", 11)).grid(row=0, column=2, padx=5)
    price_entry = Entry(frame_top, width=10)
    price_entry.grid(row=0, column=3, padx=5)
    Label(frame_top, text="الكمية:", bg="#F8F6F0", font=("Tajawal", 11)).grid(row=0, column=4, padx=5)
    qty_entry = Entry(frame_top, width=10)
    qty_entry.grid(row=0, column=5, padx=5)
    Label(frame_top, text="الفئة:", bg="#F8F6F0", font=("Tajawal", 11)).grid(row=0, column=6, padx=5)
    cat_entry = Entry(frame_top, width=12)
    cat_entry.grid(row=0, column=7, padx=5)
    img_path_var = StringVar()

    def choose_image():
        path = filedialog.askopenfilename(filetypes=[("PNG Images","*.png"), ("JPEG Images","*.jpg;*.jpeg")])
        if path:
            img_path_var.set(path)
            messagebox.showinfo("تم", f"تم اختيار الصورة:\n{path}")

    Button(frame_top, text="📸 اختيار صورة", bg="#EDDBC0", command=choose_image).grid(row=0, column=8, padx=5)

    # ====== الجدول ======
    frame_table = Frame(win)
    frame_table.pack(pady=10, fill=BOTH, expand=True)
    scroll_y = Scrollbar(frame_table, orient=VERTICAL)
    scroll_y.pack(side=RIGHT, fill=Y)

    style = ttk.Style()
    style.configure("Treeview", font=("Tajawal", 11), rowheight=26)
    style.configure("Treeview.Heading", font=("Tajawal", 12, "bold"))

    table = ttk.Treeview(frame_table, columns=("1", "2", "3", "4", "5"), show="headings", yscrollcommand=scroll_y.set)
    scroll_y.config(command=table.yview)
    table.pack(fill=BOTH, expand=True)
    table.heading("1", text="المنتج")
    table.heading("2", text="السعر")
    table.heading("3", text="الكمية")
    table.heading("4", text="الفئة")
    table.heading("5", text="الصورة")
    table.column("1", width=200, anchor="center")
    table.column("2", width=100, anchor="center")
    table.column("3", width=100, anchor="center")
    table.column("4", width=150, anchor="center")
    table.column("5", width=200, anchor="center")

    def load_products(filter_text=""):
        table.delete(*table.get_children())
        for row in ws.iter_rows(min_row=2, values_only=True):
            if filter_text.lower() in str(row[0]).lower() or filter_text.lower() in str(row[3]).lower():
                item = table.insert("", END, values=row)
                try:
                    qty = int(row[2])
                    if qty < 5:
                        table.item(item, tags=("low",))
                except: pass
        table.tag_configure("low", background="#FFCCCC")

    def add_product():
        name = name_entry.get().strip()
        price = price_entry.get().strip()
        qty = qty_entry.get().strip()
        cat = cat_entry.get().strip()
        img_path = img_path_var.get().strip()
        if not name or not price or not qty:
            messagebox.showwarning("تنبيه", "الرجاء إدخال جميع الحقول المطلوبة.")
            return
        try: float(price); int(qty)
        except: messagebox.showwarning("خطأ", "يرجى إدخال أرقام صحيحة للسعر والكمية."); return
        ws.append([name, price, qty, cat, img_path])
        wb.save("products.xlsx")
        messagebox.showinfo("تم", f"✅ تمت إضافة المنتج ({name}) بنجاح.")
        name_entry.delete(0, END); price_entry.delete(0, END)
        qty_entry.delete(0, END); cat_entry.delete(0, END)
        img_path_var.set("")
        load_products()

    Button(frame_top, text="➕ إضافة", bg="#6D8B74", fg="white", width=12, command=add_product).grid(row=2, column=2, pady=10)
    load_products()

# ====== تشغيل البرنامج من شاشة الدخول ======
login_screen()
