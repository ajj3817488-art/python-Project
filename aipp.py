from tkinter import *
from tkinter import ttk, filedialog, messagebox
import datetime
import openpyxl
from openpyxl import Workbook
from zipfile import BadZipFile
import shutil
import csv
import os

# =================== إعداد النافذة الرئيسية ===================
SA = Tk()
SA.geometry("1200x550")
SA.title("Market tools for Building [ملابس]")
SA.configure(bg="#EFEAD8")

# التاريخ الحالي
mae = datetime.datetime.now()
date = mae.strftime("%Y-%m-%d")

# =================== إنشاء ملفات Excel إذا لم تكن موجودة ===================
def create_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "customer"
    ws.append(["Full Name", "Phone", "Address", "Total", "Date"])
    wb.save("raken.xlsx")
    
def create_products_excel():
    try:
        wb = openpyxl.load_workbook("products.xlsx")
    except (FileNotFoundError, BadZipFile):
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Product Name", "Price", "Quantity", "Category"])
        wb.save("products.xlsx")

# تحقق من وجود الملفات
try:
    wb = openpyxl.load_workbook("raken.xlsx")
except (FileNotFoundError, BadZipFile):
    create_excel()

create_products_excel()

# =================== النسخ الاحتياطي ===================
def export_backup():
    def export_excel():
        try:
            source = "raken.xlsx"
            backup_name = f"backup_raken_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            dest_path = filedialog.asksaveasfilename(initialfile=backup_name,
                                                     defaultextension=".xlsx",
                                                     filetypes=[("Excel Files", "*.xlsx")])
            if dest_path:
                shutil.copy(source, dest_path)
                messagebox.showinfo("✅ تم", f"تم حفظ النسخة الاحتياطية (Excel) بنجاح:\n{dest_path}")
            win.destroy()
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء التصدير:\n{e}")

    def export_csv():
        try:
            wb = openpyxl.load_workbook("raken.xlsx")
            ws = wb.active
            backup_name = f"backup_raken_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            dest_path = filedialog.asksaveasfilename(initialfile=backup_name,
                                                     defaultextension=".csv",
                                                     filetypes=[("CSV Files", "*.csv")])
            if dest_path:
                with open(dest_path, "w", newline="", encoding="utf-8-sig") as file:
                    writer = csv.writer(file)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(row)
                messagebox.showinfo("✅ تم", f"تم حفظ النسخة الاحتياطية (CSV) بنجاح:\n{dest_path}")
            win.destroy()
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء إنشاء CSV:\n{e}")

    win = Toplevel(SA)
    win.title("💾 تصدير نسخة احتياطية")
    win.geometry("350x180")
    win.resizable(False, False)
    win.configure(bg="#EFEAD8")

    Label(win, text="اختر نوع النسخة الاحتياطية:", bg="#EFEAD8", font=("Tajawal", 12, "bold")).pack(pady=15)
    Button(win, text="📘 Excel (.xlsx)", bg="#6D8B74", fg="white", width=20, font=("Tajawal", 11), command=export_excel).pack(pady=5)
    Button(win, text="📄 CSV (.csv)", bg="#6D8B74", fg="white", width=20, font=("Tajawal", 11), command=export_csv).pack(pady=5)

# =================== حفظ الفاتورة ===================
def save():
    uzr = F2.get().strip()
    AM = F3.get().strip()
    anon = F4.get().strip()
    AG = F6.get().strip()
    JH = F8.get().strip()

    if uzr == "" or AM == "":
        messagebox.showwarning("تنبيه", "❌ الرجاء إدخال اسم المشتري ورقم الهاتف.")
        return

    try:
        Ayman = openpyxl.load_workbook("raken.xlsx")
    except (FileNotFoundError, BadZipFile):
        create_excel()
        Ayman = openpyxl.load_workbook("raken.xlsx")

    flk = Ayman.active
    flk.append([uzr, AM, anon, AG, JH])
    Ayman.save("raken.xlsx")
    Ayman.close()
    messagebox.showinfo("تم", "✅ تم حفظ الفاتورة بنجاح!")
    clear1()

# =================== إدارة المنتجات ===================
def manage_products():
    win = Toplevel(SA)
    win.title("📦 إدارة المنتجات")
    win.geometry("800x400")
    win.resizable(False, False)
    win.configure(bg="#EFEAD8")

    wb = openpyxl.load_workbook("products.xlsx")
    ws = wb.active

    frame_top = Frame(win, bg="#EFEAD8")
    frame_top.pack(pady=10)

    Label(frame_top, text="اسم المنتج:", bg="#EFEAD8").grid(row=0, column=0, padx=5)
    name_entry = Entry(frame_top, width=15)
    name_entry.grid(row=0, column=1, padx=5)

    Label(frame_top, text="السعر:", bg="#EFEAD8").grid(row=0, column=2, padx=5)
    price_entry = Entry(frame_top, width=10)
    price_entry.grid(row=0, column=3, padx=5)

    Label(frame_top, text="الكمية:", bg="#EFEAD8").grid(row=0, column=4, padx=5)
    qty_entry = Entry(frame_top, width=10)
    qty_entry.grid(row=0, column=5, padx=5)

    Label(frame_top, text="الفئة:", bg="#EFEAD8").grid(row=0, column=6, padx=5)
    cat_entry = Entry(frame_top, width=12)
    cat_entry.grid(row=0, column=7, padx=5)

    # جدول المنتجات
    frame_table = Frame(win)
    frame_table.pack(pady=10, fill=BOTH, expand=True)

    scroll_y = Scrollbar(frame_table, orient=VERTICAL)
    scroll_y.pack(side=RIGHT, fill=Y)

    table = ttk.Treeview(frame_table, columns=("1", "2", "3", "4"), show="headings", yscrollcommand=scroll_y.set)
    scroll_y.config(command=table.yview)
    table.pack(fill=BOTH, expand=True)

    table.heading("1", text="المنتج")
    table.heading("2", text="السعر")
    table.heading("3", text="الكمية")
    table.heading("4", text="الفئة")

    table.column("1", width=200, anchor="center")
    table.column("2", width=100, anchor="center")
    table.column("3", width=100, anchor="center")
    table.column("4", width=150, anchor="center")

    # =================== وظائف المنتجات ===================
    def load_products():
        table.delete(*table.get_children())
        for row in ws.iter_rows(min_row=2, values_only=True):
            table.insert("", END, values=row)

    def add_product():
        name = name_entry.get().strip()
        price = price_entry.get().strip()
        qty = qty_entry.get().strip()
        cat = cat_entry.get().strip()

        if not name or not price or not qty:
            messagebox.showwarning("تنبيه", "الرجاء إدخال جميع الحقول المطلوبة.")
            return
        try:
            float(price)
            int(qty)
        except ValueError:
            messagebox.showwarning("خطأ", "يرجى إدخال أرقام صحيحة للسعر والكمية.")
            return

        ws.append([name, price, qty, cat])
        wb.save("products.xlsx")
        wb.close()
        messagebox.showinfo("تم", f"✅ تمت إضافة المنتج ({name}) بنجاح.")
        name_entry.delete(0, END)
        price_entry.delete(0, END)
        qty_entry.delete(0, END)
        cat_entry.delete(0, END)
        # إعادة تحميل الجدول
        wb_re = openpyxl.load_workbook("products.xlsx")
        ws_re = wb_re.active
        nonlocal ws, wb
        ws = ws_re
        wb = wb_re
        load_products()

    def delete_product():
        selected = table.selection()
        if not selected:
            messagebox.showwarning("تنبيه", "الرجاء اختيار منتج للحذف.")
            return
        item = table.item(selected[0])["values"][0]
        confirm = messagebox.askyesno("تأكيد", f"هل تريد حذف المنتج '{item}'؟")
        if confirm:
            data = list(ws.iter_rows(min_row=2, values_only=True))
            ws.delete_rows(2, ws.max_row)
            for r in data:
                if r[0] != item:
                    ws.append(r)
            wb.save("products.xlsx")
            wb.close()
            # إعادة تحميل الجدول والبيانات
            wb_re = openpyxl.load_workbook("products.xlsx")
            ws_re = wb_re.active
            nonlocal ws, wb
            ws = ws_re
            wb = wb_re
            load_products()
            messagebox.showinfo("تم", "✅ تم حذف المنتج بنجاح.")

    def edit_product():
        selected = table.selection()
        if not selected:
            messagebox.showwarning("تنبيه", "الرجاء اختيار منتج للتعديل.")
            return
        values = table.item(selected[0])["values"]
        name_entry.delete(0, END)
        price_entry.delete(0, END)
        qty_entry.delete(0, END)
        cat_entry.delete(0, END)
        name_entry.insert(0, values[0])
        price_entry.insert(0, values[1])
        qty_entry.insert(0, values[2])
        cat_entry.insert(0, values[3])

        def save_edit():
            new_name = name_entry.get().strip()
            new_price = price_entry.get().strip()
            new_qty = qty_entry.get().strip()
            new_cat = cat_entry.get().strip()
            if not new_name or not new_price or not new_qty:
                messagebox.showwarning("تنبيه", "الرجاء إدخال جميع الحقول المطلوبة.")
                return
            try:
                float(new_price)
                int(new_qty)
            except ValueError:
                messagebox.showwarning("خطأ", "يرجى إدخال أرقام صحيحة للسعر والكمية.")
                return

            # تحديث الصف في Excel
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                if row[0].value == values[0]:
                    ws.cell(i,1).value = new_name
                    ws.cell(i,2).value = new_price
                    ws.cell(i,3).value = new_qty
                    ws.cell(i,4).value = new_cat
                    break
            wb.save("products.xlsx")
            wb.close()
            # إعادة تحميل الجدول
            wb_re = openpyxl.load_workbook("products.xlsx")
            ws_re = wb_re.active
            nonlocal ws, wb
            ws = ws_re
            wb = wb_re
            load_products()
            messagebox.showinfo("تم", f"✅ تم تعديل المنتج ({new_name}) بنجاح.")
            name_entry.delete(0, END)
            price_entry.delete(0, END)
            qty_entry.delete(0, END)
            cat_entry.delete(0, END)
            edit_btn.config(state=NORMAL)
            add_btn.config(state=NORMAL)
            save_edit_btn.destroy()

        add_btn.config(state=DISABLED)
        edit_btn.config(state=DISABLED)
        save_edit_btn = Button(frame_top, text="💾 حفظ التعديل", bg="#6D8B74", fg="white", width=15, command=save_edit)
        save_edit_btn.grid(row=1, column=1, pady=10)

    # الأزرار
    add_btn = Button(frame_top, text="➕ إضافة منتج", bg="#6D8B74", fg="white", width=15, command=add_product)
    add_btn.grid(row=1, column=0, pady=10)
    edit_btn = Button(frame_top, text="✏ تعديل المنتج", bg="#5A9BD5", fg="white", width=15, command=edit_product)
    edit_btn.grid(row=1, column=1, pady=10)
    delete_btn = Button(frame_top, text="🗑 حذف المنتج", bg="#C65D7B", fg="white", width=15, command=delete_product)
    delete_btn.grid(row=1, column=2, pady=10)

    load_products()

# =================== عرض جميع الفواتير ===================
def show_all_invoices():
    try:
        wb = openpyxl.load_workbook("raken.xlsx")
        ws = wb.active
    except (FileNotFoundError, BadZipFile):
        messagebox.showerror("خطأ", "الملف 'raken.xlsx' غير موجود أو تالف!")
        return

    win = Toplevel(SA)
    win.title("📋 جميع الفواتير المسجلة")
    win.geometry("750x400")
    win.resizable(False, False)
    win.configure(bg="#EFEAD8")

    Label(win, text="📋 قائمة الفواتير المسجلة", bg="#EFEAD8", font=("Tajawal", 13, "bold")).pack(pady=10)

    frame = Frame(win)
    frame.pack(padx=10, pady=10, fill=BOTH, expand=True)

    scroll_y = Scrollbar(frame, orient=VERTICAL)
    scroll_y.pack(side=RIGHT, fill=Y)

    table = ttk.Treeview(frame, columns=("1", "2", "3", "4", "5"), show="headings", yscrollcommand=scroll_y.set)
    scroll_y.config(command=table.yview)
    table.pack(fill=BOTH, expand=True)

    table.heading("1", text="الاسم")
    table.heading("2", text="الهاتف")
    table.heading("3", text="العنوان")
    table.heading("4", text="الإجمالي")
    table.heading("5", text="التاريخ")

    table.column("1", width=150, anchor="center")
    table.column("2", width=100, anchor="center")
    table.column("3", width=200, anchor="center")
    table.column("4", width=80, anchor="center")
    table.column("5", width=100, anchor="center")

    for row in ws.iter_rows(min_row=2, values_only=True):
        table.insert("", END, values=row)

# =================== البحث عن مشتري ===================
def search_customer():
    def do_search():
        key = entry_search.get().strip()
        if not key:
            messagebox.showwarning("تنبيه", "الرجاء إدخال اسم المشتري أو رقم الهاتف.")
            return
        try:
            wb = openpyxl.load_workbook("raken.xlsx")
            ws = wb.active
        except (FileNotFoundError, BadZipFile):
            messagebox.showerror("خطأ", "📁 الملف 'raken.xlsx' غير صالح أو مفقود.")
            return
        results = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, phone, address, total, date = row
            if key.lower() in str(name).lower() or key in str(phone):
                results.append(row)
        listbox.delete(0, END)
        if results:
            for r in results:
                info = f"الاسم: {r[0]} | الهاتف: {r[1]} | العنوان: {r[2]} | المجموع: {r[3]} | التاريخ: {r[4]}"
                listbox.insert(END, info)
        else:
            listbox.insert(END, "❌ لا توجد نتائج مطابقة.")

    win = Toplevel(SA)
    win.title("🔍 البحث عن مشتري")
    win.geometry("600x400")
    win.resizable(False, False)
    win.configure(bg="#EFEAD8")

    Label(win, text="🔎 أدخل اسم المشتري أو رقم الهاتف:", bg="#EFEAD8", font=("Tajawal", 12)).pack(pady=10)
    entry_search = Entry(win, font=("Tajawal", 12), width=40, justify=CENTER)
    entry_search.pack(pady=5)
    Button(win, text="بحث", width=15, bg="#6D8B74", fg="white", font=("Tajawal", 11), command=do_search).pack(pady=5)
    listbox = Listbox(win, width=80, height=15, font=("Tajawal", 11))
    listbox.pack(pady=10)

# =================== الواجهة الرئيسية للمتجر ===================
sa = Frame(SA, bg='silver', width=600, height=500)
sa.place(x=1, y=1)

menu = {0: ["فنيلة", 30], 1: ["قميص", 59], 2: ["شرت", 20]}
F2 = F3 = F4 = F6 = F8 = None

# جدول الفواتير في الواجهة
SA1 = Frame(SA, bg="gray", width=343, height=550)
SA1.place(x=604, y=1)
hj = ttk.Treeview(SA1, selectmode="browse")
hj.place(x=1, y=1, width=340, height=550)
hj["columns"] = ("1", "2", "3")
hj.column("#0", width=80, anchor="c")
hj.column("1", width=50, anchor="c")
hj.column("2", width=50, anchor="c")
hj.column("3", width=60, anchor="c")
hj.heading("#0", text="المواد", anchor="c")
hj.heading("1", text="السعر", anchor="c")
hj.heading("2", text="العدد", anchor="c")
hj.heading("3", text="الإجمالي", anchor="c")

# =================== دوال الفاتورة ===================
def bill():
    global F2, F3, F4, F6, F8
    SA.geometry("1200x550")
    F1 = Frame(SA, bg="#5F7161", width=250, height=500, bd=2, relief=GROOVE)
    F1.place(x=950, y=1)

    Label(F1, text="اسم الم
