#nam=input("olkm")
#print("nam" +nam)
#ag=int(input())
#print("nam"+str(ag))
#gh=float(input("hjfh"))
# #print("dfh "+ str(gh))
# print("n\اهلن بك")
# man1=float(input("ادخل الرقم:"))
# nam2=float(input("ادخل الرقم الاخر"))
# print("-------")

#tr(man1+nam2))
#str (man1-nam2))
#"
# print("ضرب"  +  str(man1*nam2))
# nam1=float(input("ادخل الرقمد" ))
# # gmk=input("اختار بين (/,*و-و+و%)")
# nam2=float(input("ادخل الرقم التالي:" ))
# if gmk == "+":
#     aym=nam1+nam2
#     print(nam1, "+" , nam2 , "= " , aym)
# elif gmk=="-":
#     aym=nam1-nam2
#     print(nam1, "-"  ,nam2, "= " , aym)
       
# elif gmk=="%":
#     aym=nam1%nam2
#     print(nam1 ,"%" , nam2, "=" , aym)
# elif gmk=="/":
#     aym=nam1/nam2 nac= filter(input("hlo")) 
# maq=float(input("hlo"))
# gka=input("*,%,+,-,/")
# maq1=float(input("nol"))
# if gka=="*":
#     x= maq*maq1
#     print(maq,"*",maq1,"=",x)
# elif gka=="%":
#     x= maq%maq1
#     print(maq,"%",maq1,"=",x)
# elif gka=="+":
#     x= maq+maq1
#     print(maq,"+",maq1,"=",x)
# elif gka=="-":
#     x= maq-maq1
    # print(maq,"-",maq1,"=",x)
# wy =["fg,jk,lk,"]
# print("brsat",wy)
# wy.remove("fg,ik")
# print("After",wy)
# e= "ayman"
# for i in range(9):
#     print("hlo")
#i=3
# while i<=11:
#     print(i)
#     i+=2
# L="ayman"
# for x in range(10):
#     print(L) 

# print("حاسبه")
# man2=float(input("رقم"))
# man1=input("%,*,+,-")
# man12=float(input("رقم"))

# print("_______")

# if man1=="*":
#     c= man12*man2
#     print(man2,"*",man12,"=",c)
# elif man1=="+":
#     c=man2+man12
#     print(man12,"+",man2,"=",c)
# elif man1=="-":
#     c=man12-man2
#     print(man2,"-",man12,"c")
# elif man1=="%":
#     c=man2%man2
#     print(man12,"%",man2,"=",c)
# # else :
# print("اتفضل اطلب")
# uazr=input("sey,koi,napyd")
# uazr2=float(input("56,30,69"))
# if uazr=="sey":
#     print("yas")
#     f=input("$,=")
#     if f=="$":
#      print("دفع كاش",f)
#     elif f=="=":
#      print("ne",f)
# else:

#     print("غلط")

# if uazr=="koi":
#    w= input("klm,nam") 
#    if w=="klm":
#     print("متاح",w)



# if uazr=="napyd":
   
#    bd =input("j,g")
#    if bd=="j":
#      print("مخنوث",bd) 
#    elif bd=="g":
#       print("shgf")
#    else:
#       print("rt") 
# if uazr2=="30":
#    qw=float(input("12,13,76"))
#    if qw=="12":
#       print("kh[p]",qw)
#    elif qw=="13":
#       print("راسب",qw)









# from tkinter import *
# import tkinter as tk
# mast= tk.Tk()
# mast.geometry("400x500")
# mast.title("Ayman App")
# mast.configure(bg="red")

# label = tk.Label(mast, text="whats name :")
# label.pack()
# import random
# print("olkm👌")
# asd=random.randint(1,10 )
# xwa=0
# while True:
#      was=int(input("1 and 10 :")) 
#      was=int(was)
#      xwa+=1
#      if asd==was:
#           print(f"yaaaas in {xwa} محاوله😎")
#           break
          
          
#      elif asd<was:
#            print("اصغر↓↺")
#      elif asd>was:
#                print("اكبر↑↺")
#      else:
#             print("no")

# mast.mainloop()

# import tkinter
# from tkinter import*

# root=Tk()
# root.title("حاسبه ")
# root.geometry("560x600+100+200")
# root.resizable(width= False,height=False)
# root.configure(bg="#35353F")
# lab_k=Label(root,width=25,height=2,text="",font=("arial",30))


# may=""


# def show(value):

#   global may
#   may = may+value
#   lab_k.config(text=may)
  
# def clear () : 
  
#   global may
#   may=""
#   lab_k.config(text=may) 

# def calculate():
#   global may
#   result=""
#   if may !="":
#      try:
  
#       result=eval(may)
#      except:
#       result="غلط يبني"
#      may=""
#      lab_k.config(text=result) 
                



    







# lab_k.pack()
# Button(root, text="C", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#3697f5", command=lambda: clear() ).place(x=10 , y=100)
# Button(root, text="%", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000000", command=lambda: show("%") ).place(x=150 , y=100)
# Button(root, text="/", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000000", command=lambda: show("/") ).place(x=290 , y=100)
# Button(root, text="*", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000000", command=lambda: show("*") ).place(x=430 , y=100)

# Button(root, text="7", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000000", command=lambda: show("7") ).place(x=10 , y=200)
# Button(root, text="8", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000000", command=lambda: show("8") ).place(x=150 , y=200)
# Button(root, text="9", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000000", command=lambda: show("9") ).place(x=290 , y=200)
# Button(root, text="-", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000000", command=lambda: show("-") ).place(x=430 , y=200)

# Button(root, text="4", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000000", command=lambda: show("4") ).place(x=10 , y=300)
# Button(root, text="5", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000000", command=lambda: show("5") ).place(x=150 , y=300)
# Button(root, text="6", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000000", command=lambda: show("6") ).place(x=290 , y=300)
# Button(root, text="+", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000102", command=lambda: show("+") ).place(x=430 , y=300)

# Button(root, text="3", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000000", command=lambda: show("3") ).place(x=10 , y=400)
# Button(root, text="2", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000000", command=lambda: show("2") ).place(x=150 , y=400)
# Button(root, text="1", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000000", command=lambda: show("1") ).place(x=290 , y=400)
# Button(root, text="$", width=5, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000102", command=lambda: show("$") ).place(x=430 , y=400)
# Button(root, text="0", width=10, height=1, font=("arial" ,35,"bold"), bd=1, fg="#fff",bg="#000001", command=lambda: show("0") ).place(x=10 , y=500)
# Button(root, text=".", width=3, height=1, font=("arial",30,"bold"), bd=1, fg="#fff",bg="#000001", command=lambda: show(".") ).place(x=315 , y=500)
# Button(root, text="=", width=5, height=1, font=("arial",40,"bold"), bd=1, fg="#fff",bg="#106F3F", command=lambda: calculate() ).place(x=400, y=500)


# root.mainloop()

   
























#     print("mnjh")
#     input('hkj')
#     print('kj')
# uae=[57,24,95,70,40,56,11]
# gh=[]
# for x in uae:
#     if x in uae:
#         hj=x+5
#         gh.append(hj)
# print("hgjs",uae)
# print("uy",gh)
     #  print("fghdjs",gh)




# uazr=input("ادخل البريد")
# passord=input("كلمه السر")
# if uazr=="ayman"and passord==1234:
#     print("ys")
# elif uazr=="ayman"or passord==1234:
#     print("no")
# else:
#     print("لايوجد")

# print("hlo")
# man1=input("kl,no,hg")
# if man1=="no":
#     print("yas")
# elif man1=="kl":
#     x=input("d,k,g")
#     if x=="g":
#         print("متوفر")
#     elif x=="k":
#         print("غير")
#     elif x=="d":
#         print("neo")
# elif man1=="hg":
#         z=float(input("7,8,1"))
#         if z==7:
#             print("سبافون")
#         elif z==8:
#             print("ام فلوس")
#         elif z==1:
#             print("الشرطه")
        
#         else:
#             print("غلط")
           
from tkinter import *
from tkinter import ttk
import datetime
import openpyxl
from openpyxl import Workbook
from tkinter import messagebox, Toplevel

# إنشاء النافذة الرئيسية
SA = Tk()
SA.geometry("950x550")
SA.title("Market tools for Building [ملابس]")

# التاريخ الحالي
mae = datetime.datetime.now()
date = mae.strftime("%Y-%m-%d")

# إنشاء ملف Excel إن لم يكن موجود
try:
    wb = openpyxl.load_workbook("raken.xlsx")
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.title = "customer"
    ws.append(["Full Name", "Phone", "Address", "Total", "Date"])
    wb.save("raken.xlsx")

# دالة حفظ الفاتورة
def save():
    uzr = F2.get().strip()
    AM = F3.get().strip()
    anon = F4.get().strip()
    AG = F6.get().strip()
    JH = F8.get().strip()

    if uzr == "" or AM == "":
        messagebox.showwarning("تنبيه", "❌ الرجاء إدخال اسم المشتري ورقم الهاتف.")
        return

    Ayman = openpyxl.load_workbook("raken.xlsx")
    flk = Ayman.active
    flk.append([uzr, AM, anon, AG, JH])
    Ayman.save("raken.xlsx")
    messagebox.showinfo("تم", "✅ تم حفظ الفاتورة بنجاح!")

# دالة حذف فاتورة
def delete_invoice():
    def do_delete():
        key = entry_delete.get().strip()
        if not key:
            messagebox.showwarning("تنبيه", "يرجى إدخال الاسم أو رقم الهاتف.")
            return

        try:
            wb = openpyxl.load_workbook("raken.xlsx")
            ws = wb.active
        except FileNotFoundError:
            messagebox.showerror("خطأ", "الملف 'raken.xlsx' غير موجود!")
            return

        deleted = False
        rows_to_keep = []
        for row in ws.iter_rows(values_only=True):
            if row[0] == "Full Name":
                rows_to_keep.append(row)
                continue
            name, phone = row[0], row[1]
            if key.lower() in str(name).lower() or key in str(phone):
                deleted = True
            else:
                rows_to_keep.append(row)

        if deleted:
            ws.delete_rows(1, ws.max_row)
            for r in rows_to_keep:
                ws.append(r)
            wb.save("raken.xlsx")
            messagebox.showinfo("تم", "✅ تم حذف الفاتورة بنجاح.")
            win.destroy()
        else:
            messagebox.showinfo("معلومة", "❌ لم يتم العثور على أي فاتورة بهذا الاسم أو الرقم.")

    win = Toplevel(SA)
    win.title("❌ حذف فاتورة")
    win.geometry("400x200")
    win.resizable(False, False)
    win.configure(bg="#EFEAD8")

    Label(win, text="أدخل اسم المشتري أو رقم الهاتف:", bg="#EFEAD8", font=("Tajawal", 12)).pack(pady=15)
    entry_delete = Entry(win, font=("Tajawal", 12), width=30, justify=CENTER)
    entry_delete.pack(pady=5)

    Button(win, text="🗑 حذف", bg="#C65D7B", fg="white", font=("Tajawal", 11),
           width=15, command=do_delete).pack(pady=10)

# الإطارات العامة
sa = Frame(SA, bg='silver', width=600, height=500)
sa.place(x=1, y=1)

# قائمة المنتجات
menu = {
    0: ["فنيلة", 30],
    1: ["قميص", 59],
    2: ["شرت", 20],
}

# حقول الإدخال
F2 = F3 = F4 = F6 = F8 = None

def bill():
    global F2, F3, F4, F6, F8

    SA.geometry("1200x550")
    F1 = Frame(SA, bg="#5F7161", width=250, height=434, bd=2, relief=GROOVE)
    F1.place(x=950, y=1)

    Label(F1, text="اسم المشتري", bg="#5F7161", fg="white").place(x=160, y=10)
    F2 = Entry(F1, width=24, font=("Tajawal", 12), justify=CENTER)
    F2.place(x=15, y=40)

    Label(F1, text="رقم الهاتف", bg="#5F7161", fg="white").place(x=170, y=70)
    F3 = Entry(F1, width=24, font=("Tajawal", 12), justify=CENTER)
    F3.place(x=15, y=100)

    Label(F1, text="عنوان المشتري", bg="#5F7161", fg="white").place(x=160, y=130)
    F4 = Entry(F1, width=24, font=("Tajawal", 12), justify=CENTER)
    F4.place(x=15, y=160)

    Label(F1, text="الحساب الكلي", bg="#5F7161", fg="white").place(x=160, y=190)
    F6 = Entry(F1, width=24, font=("Tajawal", 12), justify=CENTER)
    F6.place(x=15, y=210)

    Label(F1, text="تاريخ الشراء", bg="#5F7161", fg="white").place(x=160, y=240)
    F8 = Entry(F1, width=24, font=("Tajawal", 12), justify=CENTER)
    F8.place(x=15, y=270)
    F8.insert(0, date)

    Button(F1, text="💾 حفظ الفاتورة", width=31, cursor="hand2",
           bg="#EDDBC0", command=save).place(x=12, y=310)

    Button(F1, text="🧹 إفراغ الحقول", width=31, cursor="hand2",
           bg="#EDDBC0", command=clear1).place(x=12, y=340)

    Button(F1, text="🔍 بحث عن مشتري", width=31, cursor="hand2",
           bg="#EDDBC0", command=search_customer).place(x=12, y=370)

    Button(F1, text="❌ حذف فاتورة", width=31, cursor="hand2",
           bg="#EDDBC0", command=delete_invoice).place(x=12, y=400)

    # إنشاء الفاتورة في الجدول
    total = 0
    hj.delete(*hj.get_children())

    for i in range(len(sb)):
        if int(sb[i].get()) > 0:
            qty = int(sb[i].get())
            price = menu[i][1]
            subtotal = qty * price
            total += subtotal
            hj.insert("", 'end', text=menu[i][0], values=(price, qty, subtotal))

    F6.insert(0, str(total) + "$")

def clear():
    """تفريغ الجدول والحقول"""
    hj.delete(*hj.get_children())
    if all([F2, F3, F4, F6, F8]):
        for field in [F2, F3, F4, F6, F8]:
            field.delete(0, END)

def clear1():
    """تفريغ الحقول فقط"""
    if all([F2, F3, F4, F6, F8]):
        for field in [F2, F3, F4, F6, F8]:
            field.delete(0, END)

def search_customer():
    """نافذة للبحث عن مشتري في ملف Excel"""
    def do_search():
        key = entry_search.get().strip()
        if not key:
            messagebox.showwarning("تنبيه", "الرجاء إدخال اسم المشتري أو رقم الهاتف.")
            return

        try:
            wb = openpyxl.load_workbook("raken.xlsx")
            ws = wb.active
        except FileNotFoundError:
            messagebox.showerror("خطأ", "الملف 'raken.xlsx' غير موجود!")
            return

        results = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, phone, address, total, date = row
            if key.lower() in str(name).lower() or key in str(phone):
                results.append(row)

        listbox.delete(0, END)
        if results:
            for r in results:
                info = f"الاسم: {r[0]} | الهاتف: {r[1]} | العنوان: {r[2]} | المجموع: {r[3]} | التاريخ: {r[4]}"
                listbox.insert(END, info)
        else:
            listbox.insert(END, "❌ لا توجد نتائج مطابقة.")

    # إنشاء نافذة البحث
    win = Toplevel(SA)
    win.title("🔍 البحث عن مشتري")
    win.geometry("600x400")
    win.resizable(False, False)
    win.configure(bg="#EFEAD8")

    Label(win, text="🔎 أدخل اسم المشتري أو رقم الهاتف:",
          bg="#EFEAD8", font=("Tajawal", 12)).pack(pady=10)
    entry_search = Entry(win, font=("Tajawal", 12), width=40, justify=CENTER)
    entry_search.pack(pady=5)

    Button(win, text="بحث", width=15, bg="#6D8B74", fg="white",
           font=("Tajawal", 11), command=do_search).pack(pady=5)

    listbox = Listbox(win, width=80, height=15, font=("Tajawal", 11))
    listbox.pack(pady=10)

# تحميل الصور
try:
    img0 = PhotoImage(file='fol/1.png')
    img1 = PhotoImage(file='fol/2.png')
    img2 = PhotoImage(file='fol/3.png')
except:
    img0 = img1 = img2 = None

# عنوان الصفحة
title = Label(sa, text="متجر الملابس", font=("Tajawal", 13),
              fg="white", bg="#5F7161", width=70)
title.place(x=0, y=0)

# الأزرار للمنتجات
man1 = Button(sa, width=88, bg="#918D7E", bd=1, relief=SOLID, cursor="hand2",
              height=85, image=img0, text="فنيلة", compound=TOP)
man1.place(x=30, y=45)

man2 = Button(sa, width=88, bg="#EFEAD8", bd=1, relief=SOLID, cursor="hand2",
              height=85, image=img1, text="شرت", compound=TOP)
man2.place(x=150, y=45)

man3 = Button(sa, width=88, bg="#EFEAD8", bd=1, relief=SOLID, cursor="hand2",
              height=85, image=img2, text="قميص", compound=TOP)
man3.place(x=290, y=45)

# صناديق الكميات
sb = []
fon = ("Times", 12, "normal")
for i in range(3):
    var = IntVar()
    spin = Spinbox(SA, from_=0, to_=5, font=fon, width=10, textvariable=var)
    spin.place(x=30 + i*120, y=140)
    sb.append(spin)

# أزرار الأوامر
Button(SA, text="🛒 شراء", fg="white", font=("Tajawal", 12),
       width=15, bg="#6D8B74", bd=1, relief=SOLID,
       cursor="hand2", height=1, command=bill).place(x=30, y=500)

Button(SA, text="🧾 فاتورة جديدة", fg="white", font=("Tajawal", 12),
       width=15, bg="#6D8B74", bd=1, relief=SOLID,
       cursor="hand2", height=1, command=clear).place(x=180, y=500)

# الجدول لعرض الفاتورة
SA1 = Frame(SA, bg="gray", width=343, height=550)
SA1.place(x=604, y=1)

hj = ttk.Treeview(SA1, selectmode="browse")
hj.place(x=1, y=1, width=340, height=550)
hj["columns"] = ("1", "2", "3")
hj.column("#0", width=80, anchor="c")
hj.column("1", width=50, anchor="c")
hj.column("2", width=50, anchor="c")
hj.column("3", width=60, anchor="c")

hj.heading("#0", text="المواد", anchor="c")
hj.heading("1", text="السعر", anchor="c")
hj.heading("2", text="العدد", anchor="c")
hj.heading("3", text="الإجمالي", anchor="c")

SA.mainloop()










