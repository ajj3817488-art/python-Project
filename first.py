import os, datetime, csv, shutil
import openpyxl
from openpyxl import Workbook
from zipfile import BadZipFile
from tkinter import *
from tkinter import ttk, messagebox, filedialog, Toplevel
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
import arabic_reshaper
from bidi.algorithm import get_display
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
# ====== نظام تسجيل الدخول ======
def create_users_excel():
    """إنشاء ملف المستخدمين إن لم يكن موجود"""
    try:
        wb = openpyxl.load_workbook("users.xlsx")
    except (FileNotFoundError, BadZipFile):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["Username", "Password", "Role"])
        ws.append(["admin", "1234", "Admin"])
        ws.append(["seller", "1111", "Seller"])
        wb.save("users.xlsx")

create_users_excel()

def login_screen():
    """نافذة تسجيل الدخول"""
    login_win = Tk()
    login_win.title("🔐 تسجيل الدخول")
    login_win.geometry("400x300")
    login_win.configure(bg="#EFEAD8")
    login_win.resizable(False, False)

    Label(login_win, text="🧾 نظام المبيعات - تسجيل الدخول", bg="#EFEAD8", font=("Tajawal", 14, "bold")).pack(pady=20)

    Label(login_win, text="اسم المستخدم:", bg="#EFEAD8").pack(pady=5)
    user_entry = Entry(login_win, width=30, font=("Tajawal", 12))
    user_entry.pack()

    Label(login_win, text="كلمة المرور:", bg="#EFEAD8").pack(pady=5)
    pass_entry = Entry(login_win, width=30, font=("Tajawal", 12), show="*")
    pass_entry.pack()

    def verify_login():
        username = user_entry.get().strip()
        password = pass_entry.get().strip()

        if username == "" or password == "":
            messagebox.showwarning("تنبيه", "الرجاء إدخال اسم المستخدم وكلمة المرور.")
            return

        try:
            wb = openpyxl.load_workbook("users.xlsx")
            ws = wb.active
        except (FileNotFoundError, BadZipFile):
            create_users_excel()
            wb = openpyxl.load_workbook("users.xlsx")
            ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            user, pw, role = row
            if username == user and password == pw:
                messagebox.showinfo("تم", f"✅ مرحبًا {username} ({role})")
                login_win.destroy()
                open_main_window(role)
                return

        messagebox.showerror("خطأ", "❌ اسم المستخدم أو كلمة المرور غير صحيحة.")

    Button(login_win, text="تسجيل الدخول", bg="#6D8B74", fg="white", width=20, font=("Tajawal", 12),
           command=verify_login).pack(pady=15)

    login_win.mainloop()

def open_main_window(role):
    """فتح البرنامج الرئيسي مع تحديد الصلاحيات"""
    global SA
    SA = Tk()
    SA.geometry("950x550")
    SA.title(f"Market tools for Building [ملابس] - {role}")

    # إذا كان المستخدم بائع، نخفي أزرار الإدارة
    if role == "Seller":
        manage_btn_state = DISABLED
        report_btn_state = DISABLED
    else:
        manage_btn_state = NORMAL
        report_btn_state = NORMAL

    # الواجهة الأساسية
    sa = Frame(SA, bg='silver', width=600, height=500)
    sa.place(x=1, y=1)

    title = Label(sa, text=f"مرحبًا {role}", font=("Tajawal", 13), fg="white", bg="#5F7161", width=70)
    title.place(x=0, y=0)

    Button(SA, text="🛒 شراء", fg="white", font=("Tajawal", 12),
           width=15, bg="#6D8B74", bd=1, relief=SOLID, cursor="hand2", height=1, command=bill).place(x=30, y=500)
    Button(SA, text="🧾 فاتورة جديدة", fg="white", font=("Tajawal", 12),
           width=15, bg="#6D8B74", bd=1, relief=SOLID, cursor="hand2", height=1, command=clear).place(x=180, y=500)
    Button(SA, text="📦 إدارة المنتجات", fg="white", font=("Tajawal", 12),
           width=15, bg="#6D8B74", bd=1, relief=SOLID, cursor="hand2", height=1,
           state=manage_btn_state, command=manage_products).place(x=330, y=500)
    Button(SA, text="📊 التقارير", fg="white", font=("Tajawal", 12),
           width=15, bg="#5F7161", bd=1, relief=SOLID, cursor="hand2", height=1,
           state=report_btn_state, command=show_dashboard).place(x=480, y=500)

    # زر تسجيل الخروج
    Button(SA, text="🚪 تسجيل الخروج", bg="#C65D7B", fg="white", font=("Tajawal", 11),
           width=15, command=lambda: (SA.destroy(), login_screen())).place(x=780, y=10)

    SA.mainloop()

# إنشاء النافذة الرئيسية


SA=Tk()   

SA.geometry("950x550")

SA.title("Market tools for Building [ملابس]")

# التاريخ الحالي
mae = datetime.datetime.now()
date = mae.strftime("%Y-%m-%d")

# دالة إنشاء ملف Excel جديد
def create_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "customer"
    ws.append(["Full Name", "Phone", "Address", "Total", "Date"])
    wb.save("raken.xlsx")

# ✅ التحقق من سلامة ملف Excel
try:
    wb = openpyxl.load_workbook("raken.xlsx")
except (FileNotFoundError, BadZipFile):
    create_excel()

# ✅ دالة لتصدير نسخة احتياطية (Excel أو CSV)
def export_backup():
    def export_excel():
        try:
            source = "raken.xlsx"
            backup_name = f"backup_raken_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            dest_path = filedialog.asksaveasfilename(
                initialfile=backup_name,
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")]
            )
            if dest_path:
                shutil.copy(source, dest_path)
                messagebox.showinfo("✅ تم", f"تم حفظ النسخة الاحتياطية (Excel) بنجاح في:\n{dest_path}")
            win.destroy()
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء التصدير:\n{e}")

    def export_csv():
        try:
            wb = openpyxl.load_workbook("raken.xlsx")
            ws = wb.active
            backup_name = f"backup_raken_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            dest_path = filedialog.asksaveasfilename(
                initialfile=backup_name,
                defaultextension=".csv",
                filetypes=[("CSV Files", "*.csv")]
            )
            if dest_path:
                with open(dest_path, "w", newline="", encoding="utf-8-sig") as file:
                    writer = csv.writer(file)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(row)
                messagebox.showinfo("✅ تم", f"تم حفظ النسخة الاحتياطية (CSV) بنجاح في:\n{dest_path}")
            win.destroy()
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء إنشاء CSV:\n{e}")

    # نافذة الخيارات
    win = Toplevel(SA)
    win.title("💾 تصدير نسخة احتياطية")
    win.geometry("350x180")
    win.resizable(False, False)
    win.configure(bg="#EFEAD8")

    Label(win, text="اختر نوع النسخة الاحتياطية:", bg="#EFEAD8", font=("Tajawal", 12, "bold")).pack(pady=15)

    Button(win, text="📘 Excel (.xlsx)", bg="#6D8B74", fg="white", width=20, font=("Tajawal", 11),
           command=export_excel).pack(pady=5)

    Button(win, text="📄 CSV (.csv)", bg="#6D8B74", fg="white", width=20, font=("Tajawal", 11),
           command=export_csv).pack(pady=5)

# دالة حفظ الفاتورة
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from bidi.algorithm import get_display
import arabic_reshaper

def arabic_text(text):
    """تجهيز النص العربي ليظهر بشكل صحيح في PDF"""
    reshaped_text = arabic_reshaper.reshape(text)
    return get_display(reshaped_text)

def save():
    uzr = F2.get().strip()
    AM = F3.get().strip()
    anon = F4.get().strip()
    AG = F6.get().strip()
    JH = F8.get().strip()

    if uzr == "" or AM == "":
        messagebox.showwarning("تنبيه", "❌ الرجاء إدخال اسم المشتري ورقم الهاتف.")
        return

    try:
        Ayman = openpyxl.load_workbook("raken.xlsx")
    except (FileNotFoundError, BadZipFile):
        create_excel()
        Ayman = openpyxl.load_workbook("raken.xlsx")

    flk = Ayman.active
    flk.append([uzr, AM, anon, AG, JH])
    Ayman.save("raken.xlsx")

    # إنشاء مجلد الفواتير إن لم يكن موجودًا
    if not os.path.exists("Invoices"):
        os.makedirs("Invoices")

    # رقم الفاتورة التلقائي
    invoice_id = flk.max_row - 1
    file_name = f"Invoices/فاتورة_{invoice_id}_{uzr}.pdf"

    # إنشاء ملف PDF
    c = canvas.Canvas(file_name, pagesize=A4)
    width, height = A4
    c.setTitle(arabic_text(f"فاتورة رقم {invoice_id}"))

    # ====== رأس الفاتورة ======
    c.setFillColor(colors.HexColor("#5F7161"))
    c.rect(0, height - 80, width, 80, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 22)
    c.drawCentredString(width / 2, height - 50, arabic_text("متجر الأدوات للبناء [ملابس]"))

    # ====== معلومات الفاتورة ======
    y = height - 120
    c.setFont("Helvetica-Bold", 13)
    c.setFillColor(colors.black)
    c.drawRightString(560, y, arabic_text(f"رقم الفاتورة: {invoice_id}"))
    c.drawRightString(200, y, arabic_text(f"التاريخ: {JH}"))
    y -= 30

    c.setFont("Helvetica", 12)
    c.drawRightString(560, y, arabic_text(f"الاسم: {uzr}"))
    c.drawRightString(200, y, arabic_text(f"الهاتف: {AM}"))
    y -= 20
    c.drawRightString(560, y, arabic_text(f"العنوان: {anon}"))

    # ====== الجدول ======
    y -= 40
    c.setFont("Helvetica-Bold", 12)
    c.drawRightString(520, y, arabic_text("المنتج"))
    c.drawRightString(400, y, arabic_text("السعر"))
    c.drawRightString(280, y, arabic_text("الكمية"))
    c.drawRightString(160, y, arabic_text("الإجمالي"))
    y -= 10
    c.line(50, y, 550, y)
    y -= 25
    c.setFont("Helvetica", 12)

    total = 0
    for i in range(len(sb)):
        qty = int(sb[i].get())
        if qty > 0:
            item_name, price = menu[i]
            subtotal = qty * price
            total += subtotal
            c.drawRightString(520, y, arabic_text(item_name))
            c.drawRightString(400, y, arabic_text(f"{price} ريال"))
            c.drawRightString(280, y, arabic_text(str(qty)))
            c.drawRightString(160, y, arabic_text(f"{subtotal} ريال"))
            y -= 20

    c.line(50, y - 5, 550, y - 5)
    y -= 30
    c.setFont("Helvetica-Bold", 14)
    c.drawRightString(550, y, arabic_text(f"الإجمالي الكلي: {total} ريال"))

    # ====== تذييل ======
    y -= 50
    c.setFont("Helvetica-Oblique", 12)
    c.drawRightString(550, y, arabic_text("شكرًا لتسوقكم معنا ❤️"))
    c.drawRightString(550, y - 15, arabic_text("متجر الأدوات للبناء - الملابس"))

    # حفظ الفاتورة
    c.save()

    clear1()
#ة المنتجات اداره======================

def create_products_excel():
    """إنشاء ملف المنتجات إذا لم يكن موجودًا"""
    try:
        wb = openpyxl.load_workbook("products.xlsx")
    except (FileNotFoundError, BadZipFile):
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Product Name", "Price", "Quantity", "Category"])
        wb.save("products.xlsx")

create_products_excel()  # تشغيلها مرة واحدة عند بدء البرنامج


# def manage_products():
#     """واجهة إدارة المنتجات"""
#     win = Toplevel(SA)
#     win.title("📦 إدارة المنتجات")
#     win.geometry("800x400")
#     win.resizable(False, False)
#     win.configure(bg="#EFEAD8")

#     # تحميل المنتجات من الملف
#     wb = openpyxl.load_workbook("products.xlsx")
#     ws = wb.active

#     # إدخال بيانات جديدة
#     frame_top = Frame(win, bg="#EFEAD8")
#     frame_top.pack(pady=10)

#     Label(frame_top, text="اسم المنتج:", bg="#EFEAD8").grid(row=0, column=0, padx=5)
#     name_entry = Entry(frame_top, width=15)
#     name_entry.grid(row=0, column=1, padx=5)

#     Label(frame_top, text="السعر:", bg="#EFEAD8").grid(row=0, column=2, padx=5)
#     price_entry = Entry(frame_top, width=10)
#     price_entry.grid(row=0, column=3, padx=5)

#     Label(frame_top, text="الكمية:", bg="#EFEAD8").grid(row=0, column=4, padx=5)
#     qty_entry = Entry(frame_top, width=10)
#     qty_entry.grid(row=0, column=5, padx=5)

#     Label(frame_top, text="الفئة:", bg="#EFEAD8").grid(row=0, column=6, padx=5)
#     cat_entry = Entry(frame_top, width=12)
    
#     cat_entry.grid(row=0, column=7, padx=5)
def manage_products():
    """📦 واجهة إدارة المنتجات مع الصور والبحث"""
    win = Toplevel(SA)
    win.title("📦 إدارة المنتجات")
    win.geometry("950x500")
    win.configure(bg="#F8F6F0")
    win.resizable(False, False)

    # ====== تحميل أو إنشاء ملف المنتجات ======
    try:
        wb = openpyxl.load_workbook("products.xlsx")
    except:
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Product Name", "Price", "Quantity", "Category", "Image Path"])
        wb.save("products.xlsx")
        wb = openpyxl.load_workbook("products.xlsx")
    ws = wb.active

    # ====== قسم الإدخال ======
    frame_top = Frame(win, bg="#F8F6F0")
    frame_top.pack(pady=10)

    Label(frame_top, text="اسم المنتج:", bg="#F8F6F0", font=("Tajawal", 11)).grid(row=0, column=0, padx=5)
    name_entry = Entry(frame_top, width=15)
    name_entry.grid(row=0, column=1, padx=5)

    Label(frame_top, text="السعر:", bg="#F8F6F0", font=("Tajawal", 11)).grid(row=0, column=2, padx=5)
    price_entry = Entry(frame_top, width=10)
    price_entry.grid(row=0, column=3, padx=5)

    Label(frame_top, text="الكمية:", bg="#F8F6F0", font=("Tajawal", 11)).grid(row=0, column=4, padx=5)
    qty_entry = Entry(frame_top, width=10)
    qty_entry.grid(row=0, column=5, padx=5)

    Label(frame_top, text="الفئة:", bg="#F8F6F0", font=("Tajawal", 11)).grid(row=0, column=6, padx=5)
    cat_entry = Entry(frame_top, width=12)
    cat_entry.grid(row=0, column=7, padx=5)

    img_path_var = StringVar()

    # ====== زر اختيار الصورة ======
    def choose_image():
        path = filedialog.askopenfilename(
            title="اختر صورة المنتج",
            filetypes=[("PNG Images", "*.png"), ("JPEG Images", "*.jpg;*.jpeg")]
        )
        if path:
            img_path_var.set(path)
            messagebox.showinfo("تم", f"تم اختيار الصورة:\n{path}")

    Button(frame_top, text="📸 اختيار صورة", bg="#EDDBC0", command=choose_image).grid(row=0, column=8, padx=5)

    # ====== حقل البحث ======
    Label(frame_top, text="🔍 بحث:", bg="#F8F6F0", font=("Tajawal", 11, "bold")).grid(row=1, column=0, pady=8)
    search_entry = Entry(frame_top, width=25, font=("Tajawal", 11))
    search_entry.grid(row=1, column=1, columnspan=3, padx=5)

    # ====== الجدول ======
    frame_table = Frame(win)
    frame_table.pack(pady=10, fill=BOTH, expand=True)

    scroll_y = Scrollbar(frame_table, orient=VERTICAL)
    scroll_y.pack(side=RIGHT, fill=Y)

    style = ttk.Style()
    style.configure("Treeview", font=("Tajawal", 11), rowheight=26)
    style.configure("Treeview.Heading", font=("Tajawal", 12, "bold"))

    table = ttk.Treeview(frame_table, columns=("1", "2", "3", "4", "5"), show="headings", yscrollcommand=scroll_y.set)
    scroll_y.config(command=table.yview)
    table.pack(fill=BOTH, expand=True)

    table.heading("1", text="المنتج")
    table.heading("2", text="السعر")
    table.heading("3", text="الكمية")
    table.heading("4", text="الفئة")
    table.heading("5", text="مسار الصورة")

    # ====== تحميل المنتجات ======
    def load_products(filter_text=""):
        table.delete(*table.get_children())
        for row in ws.iter_rows(min_row=2, values_only=True):
            if filter_text.lower() in str(row[0]).lower() or filter_text.lower() in str(row[3]).lower():
                table.insert("", END, values=row)

    # ====== إضافة منتج ======
    def add_product():
        name = name_entry.get().strip()
        price = price_entry.get().strip()
        qty = qty_entry.get().strip()
        cat = cat_entry.get().strip()
        img_path = img_path_var.get().strip()

        if not name or not price or not qty:
            messagebox.showwarning("تنبيه", "الرجاء إدخال جميع الحقول المطلوبة.")
            return
        try:
            float(price)
            int(qty)
        except ValueError:
            messagebox.showwarning("خطأ", "يرجى إدخال أرقام صحيحة للسعر والكمية.")
            return

        ws.append([name, price, qty, cat, img_path])
        wb.save("products.xlsx")
        messagebox.showinfo("تم", f"✅ تمت إضافة المنتج ({name}) بنجاح.")
        name_entry.delete(0, END)
        price_entry.delete(0, END)
        qty_entry.delete(0, END)
        cat_entry.delete(0, END)
        img_path_var.set("")
        load_products()

    # ====== حذف منتج ======
    def delete_product():
        selected = table.selection()
        if not selected:
            messagebox.showwarning("تنبيه", "الرجاء اختيار منتج للحذف.")
            return
        item = table.item(selected[0])["values"][0]
        confirm = messagebox.askyesno("تأكيد", f"هل تريد حذف المنتج '{item}'؟")
        if confirm:
            data = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] != item]
            ws.delete_rows(2, ws.max_row)
            for r in data:
                ws.append(r)
            wb.save("products.xlsx")
            load_products()
            messagebox.showinfo("تم", "✅ تم حذف المنتج بنجاح.")

    # ====== تعديل منتج ======
    def edit_product():
        selected = table.selection()
        if not selected:
            messagebox.showwarning("تنبيه", "الرجاء اختيار منتج لتعديله.")
            return

        item_values = table.item(selected[0])["values"]
        name_entry.delete(0, END); name_entry.insert(0, item_values[0])
        price_entry.delete(0, END); price_entry.insert(0, item_values[1])
        qty_entry.delete(0, END); qty_entry.insert(0, item_values[2])
        cat_entry.delete(0, END); cat_entry.insert(0, item_values[3])
        img_path_var.set(item_values[4])

        def save_edit():
            new_name = name_entry.get().strip()
            new_price = price_entry.get().strip()
            new_qty = qty_entry.get().strip()
            new_cat = cat_entry.get().strip()
            new_img = img_path_var.get().strip()

            if not new_name or not new_price or not new_qty:
                messagebox.showwarning("تنبيه", "الرجاء إدخال جميع الحقول.")
                return

            data = list(ws.iter_rows(min_row=2, values_only=True))
            ws.delete_rows(2, ws.max_row)
            for r in data:
                if r[0] == item_values[0]:
                    ws.append([new_name, new_price, new_qty, new_cat, new_img])
                else:
                    ws.append(r)
            wb.save("products.xlsx")
            load_products()
            messagebox.showinfo("تم", "✅ تم تعديل المنتج بنجاح.")

        Button(frame_top, text="💾 حفظ التعديل", bg="#5F7161", fg="white", width=15,
               command=save_edit).grid(row=2, column=6, pady=10)

    # ====== تصدير المنتجات ======
    def export_products():
        backup_name = f"products_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        dest = filedialog.asksaveasfilename(initialfile=backup_name,
                                            defaultextension=".xlsx",
                                            filetypes=[("Excel files", "*.xlsx")])
        if dest:
            shutil.copy("products.xlsx", dest)
            messagebox.showinfo("تم", f"✅ تم تصدير قائمة المنتجات إلى:\n{dest}")

    # ====== أزرار العمليات ======
    Button(frame_top, text="➕ إضافة", bg="#6D8B74", fg="white", width=12, command=add_product).grid(row=2, column=2, pady=10)
    Button(frame_top, text="✏ تعديل", bg="#5F7161", fg="white", width=12, command=edit_product).grid(row=2, column=3, pady=10)
    Button(frame_top, text="🗑 حذف", bg="#C65D7B", fg="white", width=12, command=delete_product).grid(row=2, column=4, pady=10)
    Button(frame_top, text="📤 تصدير", bg="#918D7E", fg="white", width=12, command=export_products).grid(row=2, column=5, pady=10)
    Button(frame_top, text="📤 اضافه", bg="#918D7E", fg="white", width=12, command=export_products).grid(row=2, column=6, pady=10)
    # Button(frame_top, text=" اضافه المنت", bg="#918D7E", fg="white", width=12, command=load_products_with_images).grid(row=2, column=6, pady=10)
    # ====== البحث التلقائي أثناء الكتابة ======
    def on_search(event):
        load_products(search_entry.get())
    search_entry.bind("<KeyRelease>", on_search)

    load_products()
def load_products_with_images():
    """📸 عرض المنتجات مع الصور في واجهة البيع"""
    global menu
    menu = {}

    # إزالة أي عناصر سابقة من الواجهة
    for widget in sa.winfo_children():
        widget.destroy()

    # عنوان الواجهة
    Label(sa, text="🛍️ قائمة المنتجات", font=("Tajawal", 14, "bold"),
          fg="white", bg="#5F7161", width=70, height=2).pack(pady=5)

    try:
        wb_products = openpyxl.load_workbook("products.xlsx")
        ws_products = wb_products.active
    except (FileNotFoundError, BadZipFile):
        messagebox.showerror("خطأ", "❌ لم يتم العثور على الملف 'products.xlsx'")
        return

    x = 40   # موقع أول منتج أفقيًا
    y = 70   # موقع أول منتج عموديًا
    col_count = 0

    for i, row in enumerate(ws_products.iter_rows(min_row=2, values_only=True)):
        # التعامل مع المنتجات القديمة التي لا تحتوي على عمود صورة
        if len(row) == 5:
            name, price, qty, cat, img_path = row
        elif len(row) == 4:
            name, price, qty, cat = row
            img_path = ""
        else:
            continue

        # تحميل الصورة أو إنشاء صورة افتراضية
        if img_path and os.path.exists(img_path):
            try:
                img = PhotoImage(file=img_path)
            except:
                img = PhotoImage(width=100, height=100)
                img.put(("gray",), to=(0, 0, 100, 100))
        else:
            img = PhotoImage(width=100, height=100)
            img.put(("lightgray",), to=(0, 0, 100, 100))

        # إنشاء إطار لكل منتج
        frame = Frame(sa, bg="#EFEAD8", bd=1, relief=SOLID)
        frame.place(x=x, y=y, width=140, height=180)

        # عرض الصورة
        lbl_img = Label(frame, image=img, bg="#EFEAD8")
        lbl_img.image = img  # ضروري لمنع حذف الصورة من الذاكرة
        lbl_img.pack(pady=5)

        # عرض الاسم والسعر والكمية
        Label(frame, text=name, bg="#EFEAD8", font=("Tajawal", 10, "bold")).pack()
        Label(frame, text=f"{price} ريال", bg="#EFEAD8", fg="#5F7161", font=("Tajawal", 10)).pack()
        Label(frame, text=f"المتوفر: {qty}", bg="#EFEAD8", fg="#C65D7B", font=("Tajawal", 9)).pack()

        # حفظ البيانات في القاموس
        menu[i] = [name, float(price), int(qty), cat, img_path]

        # تغيير الموقع
        x += 160
        col_count += 1
        if col_count % 5 == 0:
            x = 40
            y += 200

# def load_products_with_images():
#     """تحميل المنتجات من Excel وعرضها في واجهة البيع"""
#     global menu
#     menu = {}

#     for widget in sa.winfo_children():
#         if isinstance(widget, Button):
#             widget.destroy()

#     try:
#         wb_products = openpyxl.load_workbook("products.xlsx")
#         ws_products = wb_products.active
#     except (FileNotFoundError, BadZipFile):
#         messagebox.showerror("خطأ", "لم يتم العثور على ملف 'products.xlsx'.")
#         return

#     # تحميل المنتجات
#     x = 30
#     y = 45
#     col_count = 0

#     for i, row in enumerate(ws_products.iter_rows(min_row=2, values_only=True)):
#         # تفكيك الصف حسب طوله (بعض الصفوف القديمة ليس فيها عمود صورة)
#      if len(row) == 5:
#       name,price,qty,cat,img_path=row
#      elif len(row) == 4:
         
#        name, price, qty, cat = row
#        img_path = ""  # بدون صورة
#     else:
 


#     # تخطي الصفوف الفارغة أو غير الصحيحة


#         # تحميل الصورة
#         if img_path and os.path.exists(img_path):
#             try:
#                 img = PhotoImage(file=img_path)
#             except:
#                 img = PhotoImage(width=80, height=80)  # صورة رمادية افتراضية
#                 img.put(("gray",), to=(0, 0, 80, 80))
#         else:
#             img = PhotoImage(width=80, height=80)
#             img.put(("gray",), to=(0, 0, 80, 80))

#         # إنشاء زر المنتج
#         btn = Button(sa, width=88, height=85, bg="#EFEAD8", bd=1, relief=SOLID,
#                      text=f"{name}\n{price} ريال", image=img, compound=TOP)
#         btn.image = img  # منع حذف الصورة من الذاكرة
#         btn.place(x=x, y=y)

#         # حفظ المنتج في القائمة
#         menu[i] = [name, float(price), int(qty), cat, img_path]

#         x += 120
#         col_count += 1
#         if col_count % 5 == 0:
#             x = 30
#             y += 150


# ✅ دالة عرض جميع الفواتير
def show_all_invoices():
    try:
        wb = openpyxl.load_workbook("raken.xlsx")
        ws = wb.active
    except (FileNotFoundError, BadZipFile):
        messagebox.showerror("خطأ", "الملف 'raken.xlsx' غير موجود أو تالف!")
        return

    win = Toplevel(SA)
    win.title("📋 جميع الفواتير المسجلة")
    win.geometry("750x400")
    win.resizable(False, False)
    win.configure(bg="#EFEAD8")

    Label(win, text="📋 قائمة الفواتير المسجلة", bg="#EFEAD8", font=("Tajawal", 13, "bold")).pack(pady=10)

    frame = Frame(win)
    frame.pack(padx=10, pady=10, fill=BOTH, expand=True)

    scroll_y = Scrollbar(frame, orient=VERTICAL)
    scroll_y.pack(side=RIGHT, fill=Y)

    table = ttk.Treeview(frame, columns=("1", "2", "3", "4", "5"), show="headings", yscrollcommand=scroll_y.set)
    scroll_y.config(command=table.yview)
    table.pack(fill=BOTH, expand=True)

    table.heading("1", text="الاسم")
    table.heading("2", text="الهاتف")
    table.heading("3", text="العنوان")
    table.heading("4", text="الإجمالي")
    table.heading("5", text="التاريخ")

    table.column("1", width=150, anchor="center")
    table.column("2", width=100, anchor="center")
    table.column("3", width=200, anchor="center")
    table.column("4", width=80, anchor="center")
    table.column("5", width=100, anchor="center")

    for row in ws.iter_rows(min_row=2, values_only=True):
        table.insert("", END, values=row)
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

def show_dashboard():
    """📊 لوحة التقارير والإحصائيات"""
    try:
        wb = openpyxl.load_workbook("raken.xlsx")
        ws = wb.active
    except (FileNotFoundError, BadZipFile):
        messagebox.showerror("خطأ", "الملف 'raken.xlsx' غير موجود أو تالف!")
        return

    win = Toplevel(SA)
    win.title("📈 لوحة التقارير والإحصائيات")
    win.geometry("900x600")
    win.configure(bg="#F8F6F0")
    win.resizable(False, False)

    Label(win, text="📊 لوحة تقارير المتجر", bg="#5F7161", fg="white",
          font=("Tajawal", 16, "bold"), width=60).pack(pady=10)

    # ====== حساب الإحصائيات ======
    total_sales = 0
    total_invoices = 0
    customers = set()
    product_sales = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        name, phone, address, total, date = row
        if total:
            try:
                total_sales += float(str(total).replace("$", "").replace("ريال", "").strip())
            except:
                pass
        total_invoices += 1
        customers.add(phone)

    # ====== عرض القيم ======
    frame_stats = Frame(win, bg="#EFEAD8", bd=2, relief=GROOVE)
    frame_stats.pack(pady=20, padx=20, fill=X)

    Label(frame_stats, text=f"💵 إجمالي المبيعات: {total_sales:.2f} ريال",
          bg="#EFEAD8", font=("Tajawal", 13, "bold")).pack(pady=5)
    Label(frame_stats, text=f"🧾 عدد الفواتير: {total_invoices}",
          bg="#EFEAD8", font=("Tajawal", 13, "bold")).pack(pady=5)
    Label(frame_stats, text=f"👥 عدد العملاء: {len(customers)}",
          bg="#EFEAD8", font=("Tajawal", 13, "bold")).pack(pady=5)

    # ====== تحليل المبيعات حسب المنتج ======
    try:
        prod_wb = openpyxl.load_workbook("products.xlsx")
        prod_ws = prod_wb.active
        for row in prod_ws.iter_rows(min_row=2, values_only=True):
            name, price, qty, cat = row
            product_sales[name] = int(qty)
    except:
        pass

    if product_sales:
        fig, ax = plt.subplots(figsize=(6, 4))
        products = list(product_sales.keys())
        quantities = list(product_sales.values())

        ax.barh(products, quantities, color="#6D8B74")
        ax.set_xlabel("الكمية المتوفرة")
        ax.set_ylabel("المنتج")
        ax.set_title("🧺 الكميات المتوفرة حسب المنتج", fontweight="bold")

        for i, v in enumerate(quantities):
            ax.text(v + 0.2, i, str(v), color="black", va="center")

        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.get_tk_widget().pack(pady=15)
        canvas.draw()
    else:
        Label(win, text="❌ لا توجد بيانات منتجات للعرض.", bg="#F8F6F0",
              font=("Tajawal", 12, "italic")).pack(pady=20)

    # ====== زر الإغلاق ======
    Button(win, text="❎ إغلاق", bg="#C65D7B", fg="white", width=15, command=win.destroy).pack(pady=15)

# واجهة البرنامج
sa = Frame(SA, bg='silver', width=600, height=500)
sa.place(x=1, y=1)

menu = {0: ["فنيلة", 30], 1: ["قميص", 59], 2: ["شرت", 20]}
F2 = F3 = F4 = F6 = F8 = None
def bill():
    global F2, F3, F4, F6, F8, tax_entry, discount_entry

    SA.geometry("1200x600")
    F1 = Frame(SA, bg="#5F7161", width=250, height=550, bd=2, relief=GROOVE)
    F1.place(x=950, y=1)

    Label(F1, text="اسم المشتري", bg="#5F7161", fg="white").place(x=160, y=10)
    F2 = Entry(F1, width=24, font=("Tajawal", 12), justify=CENTER)
    F2.place(x=15, y=40)

    Label(F1, text="رقم الهاتف", bg="#5F7161", fg="white").place(x=170, y=70)
    F3 = Entry(F1, width=24, font=("Tajawal", 12), justify=CENTER)
    F3.place(x=15, y=100)

    Label(F1, text="عنوان المشتري", bg="#5F7161", fg="white").place(x=160, y=130)
    F4 = Entry(F1, width=24, font=("Tajawal", 12), justify=CENTER)
    F4.place(x=15, y=160)

    Label(F1, text="💰 نسبة الضريبة (%)", bg="#5F7161", fg="white").place(x=120, y=190)
    tax_entry = Entry(F1, width=24, font=("Tajawal", 12), justify=CENTER)
    tax_entry.insert(0, "15")  # ضريبة افتراضية 15%
    tax_entry.place(x=15, y=210)

    Label(F1, text="💸 نسبة الخصم (%)", bg="#5F7161", fg="white").place(x=120, y=240)
    discount_entry = Entry(F1, width=24, font=("Tajawal", 12), justify=CENTER)
    discount_entry.insert(0, "0")
    discount_entry.place(x=15, y=260)

    Label(F1, text="📅 تاريخ الشراء", bg="#5F7161", fg="white").place(x=140, y=290)
    F8 = Entry(F1, width=24, font=("Tajawal", 12), justify=CENTER)
    F8.place(x=15, y=320)
    F8.insert(0, date)

    Label(F1, text="💵 الحساب النهائي", bg="#5F7161", fg="white").place(x=130, y=350)
    F6 = Entry(F1, width=24, font=("Tajawal", 12), justify=CENTER)
    F6.place(x=15, y=380)

    Button(F1, text="💾 حفظ الفاتورة", width=31, cursor="hand2",
           bg="#EDDBC0", command=save).place(x=12, y=420)
    Button(F1, text="🧹 إفراغ الحقول", width=31, cursor="hand2",
           bg="#EDDBC0", command=clear1).place(x=12, y=450)
    Button(F1, text="📊 التقارير", width=31, cursor="hand2",
           bg="#EDDBC0", command=show_dashboard).place(x=12, y=480)

    total = 0
    hj.delete(*hj.get_children())
    for i in range(len(sb)):
        if int(sb[i].get()) > 0:
            qty = int(sb[i].get())
            price = menu[i][1]
            subtotal = qty * price
            total += subtotal
            hj.insert("", 'end', text=menu[i][0], values=(f"{price} ريال", qty, f"{subtotal} ريال"))

    # حساب الضريبة والخصم
    tax_rate = float(tax_entry.get()) / 100
    discount_rate = float(discount_entry.get()) / 100
    tax_value = total * tax_rate
    discount_value = total * discount_rate
    final_total = (total + tax_value) - discount_value
    F6.insert(0, f"{final_total:.2f} ريال")

def clear():
    hj.delete(*hj.get_children())
    if all([F2, F3, F4, F6, F8]):
        for field in [F2, F3, F4, F6, F8]:
            field.delete(0, END)

def clear1():
    if all([F2, F3, F4, F6, F8]):
        for field in [F2, F3, F4, F6, F8]:
            field.delete(0, END)

def search_customer():
    def do_search():
        key = entry_search.get().strip()
        if not key:
            messagebox.showwarning("تنبيه", "الرجاء إدخال اسم المشتري أو رقم الهاتف.")
            return

        try:
            wb = openpyxl.load_workbook("raken.xlsx")
            ws = wb.active
        except (FileNotFoundError, BadZipFile):
            messagebox.showerror("خطأ", "📁 الملف 'raken.xlsx' غير صالح أو مفقود.")
            return

        results = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, phone, address, total, date = row
            if key.lower() in str(name).lower() or key in str(phone):
                results.append(row)

        listbox.delete(0, END)
        if results:
            for r in results:
                info = f"الاسم: {r[0]} | الهاتف: {r[1]} | العنوان: {r[2]} | المجموع: {r[3]} | التاريخ: {r[4]}"
                listbox.insert(END, info)
        else:
            listbox.insert(END, "❌ لا توجد نتائج مطابقة.")

    win = Toplevel(SA)
    win.title("🔍 البحث عن مشتري")
    win.geometry("600x400")
    win.resizable(False, False)
    win.configure(bg="#EFEAD8")

    Label(win, text="🔎 أدخل اسم المشتري أو رقم الهاتف:", bg="#EFEAD8", font=("Tajawal", 12)).pack(pady=10)
    entry_search = Entry(win, font=("Tajawal", 12), width=40, justify=CENTER)
    entry_search.pack(pady=5)
    Button(win, text="بحث", width=15, bg="#6D8B74", fg="white", font=("Tajawal", 11), command=do_search).pack(pady=5)
    listbox = Listbox(win, width=80, height=15, font=("Tajawal", 11))
    listbox.pack(pady=10)

# تحميل الصور (إن وجدت)
try:
    
    load_products_with_images()

    # img0 = PhotoImage(file='fol/1.png')
    # img1 = PhotoImage(file='fol/2.png')
    # img2 = PhotoImage(file='fol/3.png')
except:
       
       load_products_with_images()

    # img0 = img1 = img2 = None
def refresh_products():
    """🔄 تحديث قائمة المنتجات في واجهة البيع"""
    for widget in sa.winfo_children():
        widget.destroy()  # إزالة الأزرار القديمة
    load_products_with_images()  # إعادة تحميل المنتجات
    Button(sa, text="🔄 تحديث المنتجات", bg="#5F7161", fg="white",
           font=("Tajawal", 11), width=18, command=refresh_products).place(x=760, y=10)

title = Label(sa, text="متجر الملابس", font=("Tajawal", 13), fg="white", bg="#5F7161", width=70)
title.place(x=0, y=0)

# man1 = Button(sa, width=88, bg="#918D7E", bd=1, relief=SOLID, cursor="hand2", height=85, image=img0, text="فنيلة", compound=TOP)
# man1.place(x=30, y=45)
# man2 = Button(sa, width=88, bg="#EFEAD8", bd=1, relief=SOLID, cursor="hand2", height=85, image=img1, text="شرت", compound=TOP)
# man2.place(x=150, y=45)
# man3 = Button(sa, width=88, bg="#EFEAD8", bd=1, relief=SOLID, cursor="hand2", height=85, image=img2, text="قميص", compound=TOP)
# man3.place(x=290, y=45)

sb = []
fon = ("Times", 12, "normal")
for i in range(3):
    var = IntVar()
    spin = Spinbox(SA, from_=0, to_=5, font=fon, width=10, textvariable=var)
    spin.place(x=30 + i*120, y=140)
    sb.append(spin)

Button(SA, text="🛒 شراء", fg="white", font=("Tajawal", 12),
       width=15, bg="#6D8B74", bd=1, relief=SOLID, cursor="hand2", height=1, command=bill).place(x=30, y=500)
Button(SA, text="🧾 فاتورة جديدة", fg="white", font=("Tajawal", 12),
       width=15, bg="#6D8B74", bd=1, relief=SOLID, cursor="hand2", height=1, command=clear).place(x=180, y=500)
Button(SA, text ="قاىمه العملاء", fg="white", font=("Tajawal", 12),
       width=15, bg="#6D8B74", bd=1, relief=SOLID, cursor="hand2", height=1, command=show_dashboard).place(x=330, y=500)
Button(SA, text ="اداره المنتجات", fg="white", font=("Tajawal", 12),
       width=14, bg="#6D8B74", bd=1, relief=SOLID, cursor="hand2", height=1, command=manage_products).place(x=440, y=500)


Button(SA, text ="🔄", fg="white", font=("Tajawal", 12),
       width=6, bg="#1073C4", bd=1, relief=SOLID, cursor="hand2", height=1, command=refresh_products).place(x=500, y=460)

Button(SA, text ="اضافه المنتجات", fg="white", font=("Tajawal", 12),
       width=15, bg="#32EC89", bd=1, relief=SOLID, cursor="hand2", height=1, command=load_products_with_images).place(x=30, y=460)

SA1 = Frame(SA, bg="gray", width=343, height=550)
SA1.place(x=604, y=1)
hj = ttk.Treeview(SA1, selectmode="browse")
hj.place(x=1, y=1, width=340, height=550)
hj["columns"] = ("1", "2", "3")
hj.column("#0", width=80, anchor="c")
hj.column("1", width=50, anchor="c")
hj.column("2", width=50, anchor="c")
hj.column("3", width=60, anchor="c")
hj.heading("#0", text="المواد", anchor="c")
hj.heading("1", text="السعر", anchor="c")
hj.heading("2", text="العدد", anchor="c")
hj.heading("3", text="الإجمالي", anchor="c")

SA.mainloop()
# login_screen()
# login_screen()

# ===================== تشغيل البرنامج =====================
# if __name__ == "__main__":
#    / login_screen()

# login_screen()
